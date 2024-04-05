from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponseRedirect, Http404
from django.contrib import messages
from django.http import HttpResponse
from django.core.mail import send_mail
from django.middleware.csrf import get_token
from decimal import Decimal
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from datetime import datetime
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Count, Q
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Alignment

from cart.models import Cart, CartItem
from coupons.models import Coupon
from .forms import OrderForm, PaymentStatusForm
from .models import Order, Payment, OrderProduct, Wallet
from userprofile.models import Address
from userprofile.forms import AddressForm
from products.models import Product, Image
from .utils import render_to_pdf, send_order_confirmation_email, send_order_cancellation_email

# Create your views here.

@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='register')
def order_checkout(request):
    user            = request.user
    wallet_msg      = None
    coupon          = None
    coupon_discount = None
    wallet, create  = Wallet.objects.get_or_create(user=user)

    try:
        cart        = Cart.objects.get(user=user)
        cart_items  = CartItem.objects.filter(cart=cart)
    except Cart.DoesNotExist:
        messages.info(request, 'Your cart is empty')
        return redirect('cart')
    
    count = cart_items.count()

    if count == 0:
        messages.info(request, 'Cart is empty, add items from the shop.')
        return redirect('shop_view')
    
    for cart_item in cart_items:
        item = cart_item.product

        if not item.is_available:
            messages.info(request, f'{item.product_name} is not available right now')
            return redirect('cart')
        elif item.stock <= 0:
            messages.info(request, f'{item.product_name} is out of stock.')
            return redirect('cart')
        elif item.stock < cart_item.quantity:
            messages.info(request, f'Only {item.stock} is available for the {item.product_name}')
            return redirect('cart')
    
    products = []
    products_with_quantity = {}

    for cart_item in cart_items:
        product     = cart_item.product
        quantity    = cart_item.quantity
        products_with_quantity[product] = quantity
        
    try:
        primary_address = Address.objects.get(user=user, is_primary=True)
    except Address.DoesNotExist:
        primary_address = None
    
    cart_total = cart.cart_total
    coupon_applied = cart.coupon_name
    coupon_code  = cart.coupon_code

    total = Decimal(0)

    if coupon_code:
        try:
            coupon = Coupon.objects.get(code=coupon_code)
        except Coupon.DoesNotExist:
            messages.error(request, 'Coupon does not exist')
            return redirect('cart')
    
    price = 0
    for cart_item in cart_items:
        if cart_item.product_offer:
            total += cart_item.product_offer.apply_offer(cart_item)
        elif cart_item.category_offer:
            total += cart_item.category_offer.apply_offer(cart_item)
        else:            
            total += cart_item.product.price * cart_item.quantity
        
    if coupon:
        if coupon.discount_type == 'fixed_amount':
            coupon_discount = Decimal(coupon.discount_value).quantize(Decimal('0.00'))
        elif coupon.discount_type == 'percentage':
            coupon_discount = (total * (Decimal(coupon.discount_value) / Decimal(100))).quantize(Decimal('0.00'))
   
    grand_total = Decimal(0)
    tax         = Decimal(0)
    shipping_fee = Decimal(30)

    tax = 2 * (cart_total / 100)
    grand_total = cart_total + tax + shipping_fee

    if request.method == 'POST':
        form = OrderForm(request.POST, user=request.user)
        a_form = AddressForm()

        if form.is_valid():
            data    = Order()

            data.user           = user
            data.first_name     = form.cleaned_data['first_name']
            data.last_name      = form.cleaned_data['last_name']
            data.email          = form.cleaned_data['email']
            data.phone          = form.cleaned_data['phone']
            data.order_note     = form.cleaned_data['order_note']
            data.coupon_applied = coupon_applied
            data.coupon_discount= coupon_discount
            data.coupon_code    = coupon_code

            save_address        = request.POST.get('save-address', '')
            diff_address        = request.POST.get('diff-address', '')
            selected_address_id = form.cleaned_data['address']

            if selected_address_id:
                # selected_address = Address.objects.get(pk=selected_address_id)
                data.address        = selected_address_id
                selected_address    = str(f'{selected_address_id.address_line}, {selected_address_id.street_name}, {selected_address_id.city}, {selected_address_id.state}, {selected_address_id.country}, {selected_address_id.country}, {selected_address_id.zip_code}')
                
                data.billing_address = selected_address

            elif diff_address:
                address_line        = request.POST.get('address_line')
                street_name         = request.POST.get('street_name')
                city                = request.POST.get('city')
                state               = request.POST.get('state')
                country             = request.POST.get('country')
                zip_code            = request.POST.get('zip_code')

                data.billing_address = str(f'{address_line}, {street_name}, {city}, {state}, {country}, {zip_code}')

                if save_address:
                    new_address         = Address.objects.create(
                        user            = user,
                        address_line    = request.POST.get('address_line'),
                        street_name     = request.POST.get('street_name'),
                        city            = request.POST.get('city'),
                        state           = request.POST.get('state'),
                        country         = request.POST.get('country'),
                        zip_code        = request.POST.get('zip_code'),
                    )

                    data.address = new_address
            
            data.order_total    = grand_total
            data.tax            = tax
            data.shipping_fee   = shipping_fee
            
            payment_method = request.POST.get('payment-method', '')

            terms_and_conditions   = request.POST.get('accept-terms', '')

            if not terms_and_conditions:
                return redirect('order_checkout')
            
            for cart_item in cart_items:
                if cart_item.product.stock < cart_item.quantity:
                    messages.warning(request, f'{cart_item.product.product_name} is only {cart_item.product.stock} available.')
                    return redirect('cart')

            if payment_method == 'cash'and terms_and_conditions:
                if grand_total < 1000:
                    messages.warning(request, 'Cash on delivery is only available for orders worth morethan ₹1000/-')
                    return redirect('order_checkout')
                else:
                    payment = Payment.objects.create(
                        user            = user,
                        payment_method  = 'Cash on delivery',
                        amount          = grand_total,
                        status          = 'Pending'
                    )
                    data.payment        = payment
                    data.status         = 'Pending'
                    data.save()
                    data.order_number = f"{data.created_at.strftime('%Y%m%d')}-{data.order_id}"
                    data.save()

                    for cart_item in cart_items:
                        offer = None  # Initialize offer as None

                        # Check for ProductOffer first (assuming priority)
                        if cart_item.product_offer:
                            offer = cart_item.product_offer
                        else:
                            # Check for CategoryOffer if ProductOffer not found
                            offer = cart_item.category_offer
                        
                        product = cart_item.product
                        first_image = Image.objects.filter(product=product).first()
                        if cart_item.variation_category:
                            product_name = cart_item.product.product_name + f'({cart_item.variation_category} : {cart_item.variation_value})'
                        else:
                            product_name = cart_item.product.product_name

                        OrderProduct.objects.create(
                            order           = data,
                            product         = product,
                            product_image   = first_image.image,
                            product_name    = product_name,
                            price           = cart_item.product.price,
                            description     = cart_item.product.description,
                            category        = cart_item.product.category,
                            quantity        = cart_item.quantity,
                            offer_title     = offer.title if offer else None,
                            offer_discount_type  = offer.discount_type if offer else None,
                            offer_discount_value = offer.discount_value if offer else None,
                        )
                    
                    cart.delete()

                    return HttpResponseRedirect(reverse('order_confirmation', args=[data.order_id]))
            
            elif payment_method == 'paypal'and terms_and_conditions:
                payment = Payment.objects.create(
                    user            = user,
                    payment_method  = 'paypal',
                    amount          = grand_total,
                    status          = 'Pending'
                )
                data.payment        = payment
                data.status         = 'Pending'
                data.save()
                data.order_number = f"{data.created_at.strftime('%Y%m%d')}-{data.order_id}"
                data.save()

                for cart_item in cart_items:
                    offer = None  # Initialize offer as None

                    # Check for ProductOffer first (assuming priority)
                    if cart_item.product_offer:
                        offer = cart_item.product_offer
                    else:
                        # Check for CategoryOffer if ProductOffer not found
                        offer = cart_item.category_offer

                    product = cart_item.product
                    try:
                        first_image = Image.objects.filter(product=product).order_by('date_uploaded').first()
                    except Image.DoesNotExist:
                        first_image = None

                    if cart_item.variation_category:
                            product_name = cart_item.product.product_name + f'({cart_item.variation_category} : {cart_item.variation_value})'
                    else:
                        product_name = cart_item.product.product_name

                    OrderProduct.objects.create(
                        order           = data,
                        product         = product,
                        product_image   = first_image.image,
                        product_name    = product_name,
                        price           = cart_item.product.price,
                        description     = cart_item.product.description,
                        category        = cart_item.product.category,
                        quantity        = cart_item.quantity,
                        offer_title     = offer.title if offer else None,
                        offer_discount_type  = offer.discount_type if offer else None,
                        offer_discount_value = offer.discount_value if offer else None,
                    )

                cart.delete()

                return HttpResponseRedirect(reverse('order_confirmation', args=[data.order_id]))
            
            elif payment_method == 'Wallet'and terms_and_conditions:

                if wallet.balance >= grand_total:
                    payment = Payment.objects.create(
                        user            = user,
                        payment_method  = 'Wallet',
                        amount          = grand_total,
                        status          = 'Pending'
                    )

                    data.payment        = payment
                    data.status         = 'Pending'
                    data.save()
                    data.order_number = f"{data.created_at.strftime('%Y%m%d')}-{data.order_id}"
                    data.save()

                    for cart_item in cart_items:
                        offer = None  # Initialize offer as None

                        # Check for ProductOffer first (assuming priority)
                        if cart_item.product_offer:
                            offer = cart_item.product_offer
                        else:
                            # Check for CategoryOffer if ProductOffer not found
                            offer = cart_item.category_offer

                        product = cart_item.product
                        first_image = Image.objects.filter(product=product).first()

                        if cart_item.variation_category:
                            product_name = cart_item.product.product_name + f'({cart_item.variation_category} : {cart_item.variation_value})'
                        else:
                            product_name = cart_item.product.product_name

                        OrderProduct.objects.create(
                            order           = data,
                            product         = product,
                            product_image   = first_image.image,
                            product_name    = product_name,
                            price           = cart_item.product.price,
                            description     = cart_item.product.description,
                            category        = cart_item.product.category,
                            quantity        = cart_item.quantity,
                            offer_title     = offer.title if offer else None,
                            offer_discount_type  = offer.discount_type if offer else None,
                            offer_discount_value = offer.discount_value if offer else None,
                        )
                    
                    cart.delete()

                    return HttpResponseRedirect(reverse('order_confirmation', args=[data.order_id]))
                
                else:
                    wallet_msg = f'Your wallet balance is only ₹{wallet.balance}, so chose other options.'

    else:
        form = OrderForm(user=request.user)
        a_form = AddressForm()
    
    
    context = {
        'form': form,
        'cart_items': cart_items,
        'grand_total': grand_total,
        'a_form': a_form,
        'tax' : tax,
        'cart_total' : cart_total,
        'products' : products,
        'shippping_fee' : shipping_fee,
        'products_with_quantity': products_with_quantity,
        'csrf_token': get_token(request),
        'wallet_msg' : wallet_msg,
        'wallet' : wallet,
        'coupon_applied' : coupon_applied,
        'coupon_discount' : coupon_discount,
    }

    return render(request, 'checkout.html', context)

@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='register')
def order_confirmation(request, order_id):
    order = Order.objects.get(order_id=order_id)
  
    try:
        # Try to get the OrderProduct object using get()
        order_products = OrderProduct.objects.filter(order_id=order_id)

        if not order_products:
            raise Http404("Order product not found")
    except OrderProduct.DoesNotExist:
        # If no object is found, raise a specific error message
        raise Http404("Order product not found")

    products = [product.product for product in order_products]

    total = 0

    for order_product in order_products:
        total += order_product.quantity * order_product.product.price

    for item in products:
        if not item.is_available:
            messages.success(request, f'{item.product_name} is not available right now.')
            return redirect('order_checkout')
        
    context = {
        'order_products': order_products,
        'products': products,
        'order' : order,
        'total' : total,
    }

    payment_status = order.payment.status
    
    if payment_status == 'Completed':
        try:
            send_order_confirmation_email(order)
        except Exception:
            messages.error(request, 'Failed to send cancellation email')
        return render(request, 'order_placed.html', context)
    
    if order.status == 'Cancelled':
        messages.warning(request, 'The order has been cancelld')
        return render(request, 'order_cancelled.html', context)
        
    if request.method == 'POST':
        payment_type = order.payment.payment_method
        for order_product in order_products:
            if order_product.product.stock < order_product.quantity:
                order.status = 'Cancelled'
                order.payment.status = 'Failed'
                order.payment.save()
                order.save()
                out_of_stock_product = order_product.product

                context = {
                    'order_products': order_products,
                    'products': products,
                    'order' : order,
                    'total' : total,
                    'out_of_stock_product' : out_of_stock_product,
                }

                messages.warning(request, f'{order_product.product.product_name} is currently unavailable, Sorry for the inconvenience!')
                return render(request, 'order_cancelled.html', context)

            if payment_type == 'Cash on delivery':
                order.status = 'Accepted'
                order.is_ordered = True
                order.save()

                context = {
                    'order_products': order_products,
                    'products': products,
                    'order' : order,
                    'total' : total,
                }

                order_product.product.stock -= order_product.quantity
                order_product.product.save()

                try:
                    send_order_confirmation_email(order)
                except Exception:
                    messages.error(request, 'Failed to send cancellation email')

                return render(request, 'order_placed.html', context)
        
            elif payment_type == 'paypal':
                return HttpResponseRedirect(reverse('create_payment', args=[order.order_id]))

            elif payment_type == 'Wallet':
                wallet = Wallet.objects.get(user=request.user)
                wallet.balance -= order.order_total
                wallet.save()

                order.status = 'Accepted'
                order.is_ordered = True
                order.save()

                order.payment.status = 'Completed'
                order.payment.save()

                order_product.product.stock -= order_product.quantity
                order_product.product.save()

                context = {
                    'order_products': order_products,
                    'products': products,
                    'order' : order,
                    'total' : total,
                }

                try:
                    send_order_confirmation_email(order)
                except Exception:
                    messages.error(request, 'Failed to send cancellation email')

                return render(request, 'order_placed.html', context)
        
    context = {
        'order_products': order_products,
        'products': products,
        'order' : order,
        'total' : total,
    }

    return render(request, 'order_confirmation.html', context)


@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='register')
def payment_processing(request, order_id):
    return redirect(request, 'payment_processing.html')


@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='register')
def generate_invoice_pdf(request, order_id):
    try:
        order = Order.objects.get(order_id=order_id)
        order_products = OrderProduct.objects.filter(order=order)
    except Order.DoesNotExist:
        messages.warning(request, 'Order doesnot exist')
        return redirect('order_details')
    except OrderProduct.DoesNotExist:
         messages.warning(request, 'Order products not found')
         return redirect('order_details')
    
    context = {
        'order' : order,
        'order_products' : order_products,
    }

    return render_to_pdf('invoice_template.html', context)


@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='register')
def order_details(request, order_id):
    order = Order.objects.get(order_id=order_id)
    try:
        # Try to get the OrderProduct object using get()
        order_products = OrderProduct.objects.filter(order_id=order_id)

        if not order_products:
            raise Http404("Order product not found")
    except OrderProduct.DoesNotExist:
        # If no object is found, raise a specific error message
        raise Http404("Order product not found")
    
    context = {
        'order_products':order_products,
        'order':order,
    }

    return render(request, 'order_detail.html', context)

@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='register')
def cancel_order(request, order_id):
    user = request.user
    email = user.email
    try:
        order = get_object_or_404(Order, order_id=order_id)
        wallet = get_object_or_404(Wallet, user=user)
    except Order.DoesNotExist:
        messages.info(request, 'You dont have any orders yet.')
    except Wallet.DoesNotExist:
        messages.info(request, 'No wallet')

    if request.method == "POST":
        if order.payment.payment_method == 'Cash on delivery':
            order.status = 'Cancelled'
            order.save()
            order.payment.status = 'Cancelled'
            order.payment.save()

        elif order.payment.payment_method == 'Wallet':
            order.status = 'Cancelled'
            order.save()

            wallet.balance += order.order_total
            wallet.save()

            order.payment.status = 'refunded'
            order.payment.save()
        
        elif order.payment.payment_method == 'paypal':
            order.status = 'Cancelled'
            order.save()

            return HttpResponseRedirect(reverse('process_refund', args=[order.order_id]))

        # Send cancellation mail
        try:
            send_order_cancellation_email(order)
        except Exception:
            messages.error(request, 'Failed to send cancellation email')

        messages.success(request, 'Your order succesfully cancelled!!')
    
    return redirect('user_profile')

@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='adminlogin')
def admin_order_details(request):
    if not request.user.is_superuser:
        messages.warning(request, 'You are not authorised to view this page')
        return redirect('adminlogin') 
    
    try:
        order_product = OrderProduct.objects.all()
        orders = Order.objects.all().order_by('-created_at')
        products = []

        if order_product:
            for product in order_product:
                products.append(product)

    except Order.DoesNotExist:
        pass
    
    context = {
        'orders':orders,
        'order_product': order_product,
        'products': products,
    }

    return render(request, 'admin_order_details.html', context)

@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='adminlogin')
def admin_order_detailed_view(request, order_id):
    if not request.user.is_superuser:
        messages.warning(request, 'You are not authorised to view this page')
        return redirect('adminlogin')
    
    order = Order.objects.get(order_id=order_id)
    current_order_status = order.status
    order_status = None

    wallet = Wallet.objects.get(user=order.user)

    try:
        # Try to get the OrderProduct object using get()
        order_products = OrderProduct.objects.filter(order_id=order_id)

        if not order_products:
            raise Http404("Order product not found")
    except OrderProduct.DoesNotExist:
        # If no object is found, raise a specific error message
        raise Http404("Order product not found")

    products = [product.product for product in order_products]

    if request.method == 'POST':
        order_status = request.POST.get('order_status')
        payment_form = PaymentStatusForm(request.POST, instance=order.payment)

        if order_status:
            if order_status == 'Cancelled' and current_order_status != 'Cancelled':
                if order.payment.payment_method == 'paypal' or order.payment.payment_method == 'Wallet' and order.payment.status == 'Completed':
                    wallet.balance += order.order_total
                    wallet.save()

                    try:
                        send_mail("TechTrove - Order cancelled: ", f"Your order:{order.order_number}- for {order.full_name} is cancelled.", settings.EMAIL_HOST_USER, [order.email], fail_silently=False)
                    except Exception:
                        messages.error(request, 'Failed to send cancellation email')

            elif order_status == 'Delivered':
                messages.success(request, 'Order status changed to delivered')

            order.status = order_status
            order.save()

        if payment_form.is_valid():
            payment_form.save()
    else:
        payment_form = PaymentStatusForm(instance=order.payment)

    context = {
        'order_products': order_products,
        'products': products,
        'order' : order,
        'payment_form' : payment_form,
    }

    return render(request, 'admin_order_update.html', context)

@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='adminlogin')
def generate_sales_report(request):
    if not request.user.is_superuser:
        messages.warning(request, 'You are not authorized to this page')
        return redirect('adminlogin')
    
    if request.POST:
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        if not start_date_str or not end_date_str:
            messages.error(request, 'Both start date and end date are required')
            return redirect('admin_order_details')
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid date format')
            return redirect('admin_order_details')

        total_orders = Order.objects.filter(status='Delivered', created_at__date__range=[start_date, end_date]).count()
        total_revenue = Order.objects.filter(status='Delivered', created_at__date__range=[start_date, end_date])\
            .aggregate(Sum('order_total'))['order_total__sum'] or 0

        product_sales = OrderProduct.objects.filter(order__status='Delivered', order__created_at__date__range=[start_date, end_date])\
            .values(
                'product_name',
                'category',
                'description',
                'price',
            )\
            .annotate(
                quantity=Sum('quantity'),
                total_sales=ExpressionWrapper(
                    F('quantity') * F('price'),
                    output_field=DecimalField()
                )
            )\
            .order_by('-quantity')
        
        category_sales = OrderProduct.objects.filter(
                order__status='Delivered',
                order__created_at__date__range=[start_date, end_date]
            ).values(
                'category',
            ).annotate(
                quantity=Sum('quantity'),
            ).order_by('-quantity')

        offer_counts = OrderProduct.objects.filter(
                Q(offer_title__isnull=False) | Q(offer_discount_type__isnull=False) | Q(offer_discount_value__isnull=False)
            ) \
            .filter(order__status='Delivered', order__created_at__date__range=[start_date, end_date]) \
            .values('offer_title', 'offer_discount_value') \
            .annotate(count=Count('id'))

        coupons_applied = Order.objects.filter(
            Q(coupon_applied__isnull=False) | Q(coupon_discount__isnull=False) | Q(coupon_code__isnull=False),
            status='Delivered',
            created_at__date__range=[start_date, end_date]
        ).values('coupon_applied', 'coupon_code').annotate(overall_discount=Sum('coupon_discount')).distinct().annotate(count=Count('order_number'))  
                
        context = {
            'total_orders' : total_orders,
            'total_revenue' : total_revenue,
            'product_sales' : product_sales,
            'category_sales' : category_sales,
            'start_date' : start_date,
            'end_date' : end_date,
            'offer_counts' : offer_counts,
            'coupons_applied' : coupons_applied,
        }

    return render(request, 'sales_report.html', context)


@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='adminlogin')
def generate_sales_report_pdf(request):
    if not request.user.is_superuser:
        messages.warning(request, 'You are not authorised to view this page')
        return redirect('adminlogin')
    
    start_date_str = request.POST.get('start_date')
    end_date_str = request.POST.get('end_date')

    if not start_date_str or not end_date_str:
        messages.error(request, 'Both start date and end date are required')
        print('no dates')
        return redirect('admin_order_details')

    try:
        print(f'start_date_str: {start_date_str}, end_date_str: {end_date_str}')
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        print('Invalid date format')
        messages.error(request, 'Invalid date format')
        return redirect('admin_order_details')

    total_orders = Order.objects.filter(status='Delivered', created_at__date__range=[start_date, end_date]).count()
    total_revenue = Order.objects.filter(status='Delivered', created_at__date__range=[start_date, end_date])\
        .aggregate(Sum('order_total'))['order_total__sum'] or 0

    product_sales = OrderProduct.objects.filter(order__status='Delivered', order__created_at__date__range=[start_date, end_date])\
        .values(
            'product_name',
            'category',
            'description',
            'price',
        )\
        .annotate(
            quantity=Sum('quantity'),
            total_sales=ExpressionWrapper(
                F('quantity') * F('product__price'),
                output_field=DecimalField()
            )
        )\
        .order_by('-quantity')

    category_sales = OrderProduct.objects.filter(
            order__status='Delivered',
            order__created_at__date__range=[start_date, end_date]
        ).values(
            'category',
        ).annotate(
            quantity=Sum('quantity'),
        ).order_by('-quantity')

    offer_counts = OrderProduct.objects.filter(
            Q(offer_title__isnull=False) | Q(offer_discount_type__isnull=False) | Q(offer_discount_value__isnull=False)
        ) \
        .filter(order__status='Delivered', order__created_at__date__range=[start_date, end_date]) \
        .values('offer_title', 'offer_discount_value') \
        .annotate(count=Count('id'))

    coupons_applied = Order.objects.filter(
        Q(coupon_applied__isnull=False) | Q(coupon_discount__isnull=False) | Q(coupon_code__isnull=False),
        status='Delivered',
        created_at__date__range=[start_date, end_date]
    ).values('coupon_applied', 'coupon_code').annotate(overall_discount=Sum('coupon_discount')).distinct().annotate(count=Count('order_number'))  

    context = {
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'product_sales': product_sales,
        'category_sales': category_sales,
        'start_date': start_date,
        'end_date': end_date,
        'offer_counts': offer_counts,
        'coupons_applied': coupons_applied,
    }

    return render_to_pdf('sales_report_pdf.html', context)


@cache_control(no_cache=True, no_store=True, must_revalidate=True, max_age=0)
@login_required(login_url='adminlogin')
def generate_sales_report_excel(request):
    if not request.user.is_superuser:
        messages.warning(request, 'You are not authorised to view this page')
        return redirect('adminlogin')
    
    start_date_str = request.POST.get('start_date')
    end_date_str = request.POST.get('end_date')

    if not start_date_str or not end_date_str:
        messages.error(request, 'Both start date and end date are required')
        return redirect('admin_order_details')

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format')
        return redirect('admin_order_details')

    total_orders = Order.objects.filter(status='Delivered', created_at__date__range=[start_date, end_date]).count()
    total_revenue = Order.objects.filter(status='Delivered', created_at__date__range=[start_date, end_date])\
        .aggregate(Sum('order_total'))['order_total__sum'] or 0

    product_sales = OrderProduct.objects.filter(order__status='Delivered', order__created_at__date__range=[start_date, end_date])\
        .values(
            'product_name',
            'category',
            'description',
            'price',
        )\
        .annotate(
            quantity=Sum('quantity'),
            total_sales=ExpressionWrapper(
                F('quantity') * F('product__price'),
                output_field=DecimalField()
            )
        )\
        .order_by('-quantity')

    category_sales = OrderProduct.objects.filter(
            order__status='Delivered',
            order__created_at__date__range=[start_date, end_date]
        ).values(
            'category',
        ).annotate(
            quantity=Sum('quantity'),
        ).order_by('-quantity')

    offer_counts = OrderProduct.objects.filter(
            Q(offer_title__isnull=False) | Q(offer_discount_type__isnull=False) | Q(offer_discount_value__isnull=False)
        ) \
        .filter(order__status='Delivered', order__created_at__date__range=[start_date, end_date]) \
        .values('offer_title', 'offer_discount_value') \
        .annotate(count=Count('id'))

    coupons_applied = Order.objects.filter(
        Q(coupon_applied__isnull=False) | Q(coupon_discount__isnull=False) | Q(coupon_code__isnull=False),
        status='Delivered',
        created_at__date__range=[start_date, end_date]
    ).values('coupon_applied', 'coupon_code').annotate(overall_discount=Sum('coupon_discount')).distinct().annotate(count=Count('order_number'))

    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'Sales Report'
    worksheet.merge_cells('A1:E1')  # Merge cells for the title
    worksheet['A2'] = f'Start Date: {start_date}'
    worksheet['B2'] = f'End Date: {end_date}'

    # Define headers for product sales table
    headers_product_sales = ['Product Name', 'Category', 'Price', 'Quantity Sold', 'Revenue']

    # Write headers for product sales table
    for col_num, header in enumerate(headers_product_sales, 1):
        cell = worksheet.cell(row=5, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='left')  # Align header text to the left

    # Write data for product sales table
    for row_num, item in enumerate(product_sales, 6):
        row_data = [
            item['product_name'],
            item['category'],
            f'₹{item["price"]}',
            item['quantity'],
            f'₹{item["total_sales"]}',
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)

    # Define headers for category sales table
    headers_category_sales = ['Category Name', 'Quantity Sold']

    # Write headers for category sales table
    for col_num, header in enumerate(headers_category_sales, 1):
        cell = worksheet.cell(row=row_num + 2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='left')  # Align header text to the left

    # Write data for category sales table
    for row_num, item in enumerate(category_sales, row_num + 3):
        row_data = [
            item['category'],
            item['quantity'],
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)

    # Define headers for offer_counts table
    headers_offer_counts = ['Offer title', 'Offer discount value', 'Count']

    # Write headers for offer_counts table
    for col_num, header in enumerate(headers_offer_counts, 1):
        cell = worksheet.cell(row=row_num + 2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='left')  # Align header text to the left

    # Write data for offer_counts table
    for row_num, item in enumerate(offer_counts, row_num + 3):
        row_data = [
            item['offer_title'],
            f'₹{item["offer_discount_value"]}' if item["offer_discount_value"] else '',
            item['count'],
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)

    # Define headers for coupons_applied table
    headers_coupons_applied = ['Coupon Name', 'Overall discount', 'Count']

    # Write headers for coupons_applied table
    for col_num, header in enumerate(headers_coupons_applied, 1):
        cell = worksheet.cell(row=row_num + 2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal='left')  # Align header text to the left

    # Write data for coupons_applied table
    for row_num, item in enumerate(coupons_applied, row_num + 3):
        row_data = [
            item['coupon_applied'],
            f'₹{item["overall_discount"]}' if item["overall_discount"] else '',
            item['count'],
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)

    # Create the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
    workbook.save(response)

    return response

